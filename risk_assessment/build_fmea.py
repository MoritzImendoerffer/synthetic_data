#!/usr/bin/env python3
"""Build the post-characterization process FMEA workbook (Excel).

A-Mab-aligned methodology (post-characterization Drug-Product FMEA, Tables 5.22-
5.24; CPP rule p.212): RPN = Severity x Occurrence x Detection; a parameter is a
CPP if Severity >= 8 OR RPN > 72. The workbook shows the *initial* (pre-
characterization) RPN and the *residual* RPN after the characterization studies
and the resulting control strategy — making the risk reduction explicit — and
splits CPPs into CPP vs WC-CPP (well-controlled) per the case-study designation.

Sheets: Cover · CQA Criticality · Process FMEA (Post-PC) · Scoring Scales · Summary

Usage:  python risk_assessment/build_fmea.py
"""

from __future__ import annotations

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from amab_process import load_config                     # noqa: E402
from amab_process import studies as st                   # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "risk_assessment", "A-Mab_Post-PC_Process_Risk_Assessment.xlsx")

# --- palette (matches the report) --------------------------------------------
NAVY = "1F3864"; BLUE = "2A78D6"; LIGHT = "EAF1FB"; BAND = "F4F7FB"
RED = "F4CCCC"; AMBER = "FCE5CD"; GREEN = "D9EAD3"; GREY = "D9D9D9"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HFILL = PatternFill("solid", fgColor=NAVY)
HFONT = Font(color=WHITE, bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

# --- FMEA severity / detection maps (A-Mab-grounded) -------------------------
CQA_SEVERITY = {                       # FMEA severity (Table 5.22, 2..10)
    "afucosylation": 8, "galactosylation": 8, "high_mannose": 10, "aggregates_hmw": 8,
    "acidic_variants": 4, "hcp": 6, "residual_dna": 2, "leached_protein_a": 4,
    "lrv_xmulv": 10, "lrv_mvm": 10, "performance": 4,
}
# detection of the failure effect (Table 5.24): viral cannot be detected in DP;
# glycans/charge -> release/characterization testing; impurities -> release + IPC.
CQA_DET_RES = {"afucosylation": 6, "galactosylation": 6, "high_mannose": 6,
               "aggregates_hmw": 4, "acidic_variants": 4, "hcp": 4, "residual_dna": 4,
               "leached_protein_a": 6, "lrv_xmulv": 8, "lrv_mvm": 8, "performance": 4}

# per-parameter FMEA content: (step_key, param_key) -> dict
CONTENT = {
    ("bioreactor", "pH"): dict(cqas=["afucosylation", "galactosylation", "high_mannose", "acidic_variants", "aggregates_hmw"],
        fm="pH excursion outside the characterized range (6.6-7.1)",
        eff="Shifts glycosylation (ADCC via afucosylation; CDC via galactosylation) and charge/aggregate profile",
        ctl="Closed-loop pH control with validated/redundant probes; continuous in-process monitoring; multivariate design space 6.6-7.1"),
    ("bioreactor", "temperature"): dict(cqas=["afucosylation", "galactosylation", "acidic_variants"],
        fm="Temperature excursion outside 34-36 C",
        eff="Alters afucosylation, galactosylation and acidic charge variants; interacts with pH",
        ctl="Jacket temperature control; validated probes; in-process monitoring; design space 34-36 C"),
    ("bioreactor", "co2"): dict(cqas=["galactosylation", "acidic_variants"],
        fm="Dissolved CO2 outside 40-160 mmHg", eff="Strong effect on galactosylation and acidic variants",
        ctl="Sparge/overlay gas control; pCO2 monitoring; design space 40-160 mmHg"),
    ("bioreactor", "osmolality"): dict(cqas=["galactosylation", "acidic_variants"],
        fm="Osmolality outside 360-440 mOsm", eff="Affects galactosylation and acidic variants",
        ctl="Feed/base osmolality control; at-line monitoring; design space 360-440 mOsm"),
    ("bioreactor", "duration"): dict(cqas=["afucosylation", "galactosylation", "aggregates_hmw", "hcp", "residual_dna"],
        fm="Culture extended beyond 15-19 days / harvested on low viability",
        eff="Lowers afucosylation & galactosylation; raises HCP/DNA and aggregate as viability declines",
        ctl="Harvest criteria on viability, titer and quality; duration bounded 15-19 days"),
    ("bioreactor", "do"): dict(cqas=["performance"], fm="DO outside 30-70%",
        eff="Impacts peak VCD, titer and process consistency (no significant CQA impact)",
        ctl="Cascade DO control; in-process monitoring", kpp=True),
    ("bioreactor", "ivcc"): dict(cqas=["performance"], fm="Inoculation density outside 0.5-1.5e6/mL",
        eff="Drives peak/integral VCD and titer; no product-quality impact", ctl="Seed-train control; split-ratio spec", kpp=True),
    ("bioreactor", "medium_conc"): dict(cqas=["performance"], fm="Basal medium concentration outside 0.8-1.6X",
        eff="Statistically significant but shallow effect on afucosylation; drives titer",
        ctl="Media prep controls; concentration verified", gpp=True),
    ("bioreactor", "feed_vol"): dict(cqas=["performance"], fm="Nutrient feed-1 volume outside 9-15% WV",
        eff="Slight effect on galactosylation; critical for titer", ctl="Gravimetric feed control; feed schedule", kpp=True),
    ("protein_a", "load"): dict(cqas=["hcp"], fm="Protein load above characterized range (>50 g/L resin)",
        eff="Increases pool HCP (interaction with elution pH)", ctl="Load controlled by A280/mass balance; design space 10-50 g/L; CEX/AEX downstream HCP clearance"),
    ("protein_a", "elution_ph"): dict(cqas=["hcp"], fm="Elution pH below characterized range (<3.2)",
        eff="Increases pool HCP; also sets up viral inactivation feed", ctl="Buffer pH verified pre-use; in-line pH monitoring; design space 3.2-3.9"),
    ("protein_a", "flow"): dict(cqas=["performance"], fm="Load flow rate high with high load",
        eff="Reduces product yield (no CQA impact)", ctl="Flow controlled by pump/skid; yield IPC", kpp=True),
    ("protein_a", "end_collect"): dict(cqas=["performance"], fm="End-of-pool collection outside 2.0-3.2 CV",
        eff="Affects product yield (no CQA impact)", ctl="Collection by CV/A280; yield IPC", kpp=True),
    ("viral_inactivation", "ph"): dict(cqas=["lrv_xmulv", "aggregates_hmw"],
        fm="pH > 4.0 (incomplete inactivation) or pH < 3.2 (aggregation/precipitation)",
        eff="Loss of XMuLV inactivation assurance or increased aggregate", ctl="pH adjust to 3.5 +/- 0.1 with verified titration; in-line pH; acceptable range 3.2-4.0", cpp=True),
    ("viral_inactivation", "hold_time"): dict(cqas=["aggregates_hmw", "lrv_xmulv"],
        fm="Hold time < 60 min (incomplete) or > 180 min (aggregation)",
        eff="Incomplete inactivation or increased aggregate", ctl="Timed hold 60-120 min with alarms; CEX downstream aggregate polish"),
    ("viral_inactivation", "temperature"): dict(cqas=["lrv_xmulv"], fm="Temperature < 15 C",
        eff="Lower inactivation rate at short time / higher pH", ctl="Ambient hold 15-25 C; temperature monitored"),
    ("viral_inactivation", "protein_conc"): dict(cqas=["performance"], fm="Concentration > 35 g/L",
        eff="No significant effect on inactivation, aggregation or charge variants", ctl="Pool concentration <= 35 g/L (typically <= 20 g/L)", gpp=True),
    ("cex", "load"): dict(cqas=["aggregates_hmw", "hcp"], fm="Protein load above characterized range (>40 g/L resin)",
        eff="Reduced aggregate and HCP clearance (interactions with pH and wash conductivity)", ctl="Load by mass balance; design space 10-40 g/L; input aggregate bounded by VI"),
    ("cex", "wash_cond"): dict(cqas=["hcp"], fm="Load/wash conductivity below 3 mS/cm",
        eff="Reduced HCP clearance", ctl="Buffer conductivity verified; in-line conductivity; design space 3-7 mS/cm"),
    ("cex", "elution_ph"): dict(cqas=["aggregates_hmw"], fm="Elution pH outside 5.8-6.2",
        eff="Alters aggregate distribution and yield", ctl="Buffer pH verified; in-line pH; design space 5.8-6.2"),
    ("cex", "stop_collect"): dict(cqas=["aggregates_hmw"], fm="Extended elution stop-collect (>1.5 OD)",
        eff="Minor increase in pool aggregate", ctl="Stop-collect by descending A280; design space 0.5-1.5 OD"),
    ("cex", "flow"): dict(cqas=["performance"], fm="Elution flow rate outside 100-300 cm/hr",
        eff="Affects peak shape and yield (minor aggregate)", ctl="Flow controlled by skid", gpp=True),
    ("aex", "load_ph"): dict(cqas=["hcp", "lrv_xmulv", "lrv_mvm"], fm="Load pH below characterized range (<7.2)",
        eff="Reduced HCP removal and reduced XMuLV/MVM clearance", ctl="Load pH adjust to 7.5; in-line pH; quality design space 7.2-7.8; viral clearance validated >= 7.0"),
    ("aex", "wash1_cond"): dict(cqas=["hcp"], fm="Equil/Wash-1 conductivity above 3.6 mS/cm",
        eff="Reduced HCP removal", ctl="Buffer conductivity verified; in-line conductivity; design space 1.6-3.6 mS/cm"),
    ("aex", "load_cond"): dict(cqas=["lrv_xmulv", "lrv_mvm"], fm="Load conductivity above characterized range",
        eff="Reduced viral clearance (interacts with pH)", ctl="Conditioned load conductivity; viral clearance validated <= 15 mS/cm"),
    ("aex", "load"): dict(cqas=["hcp"], fm="Protein load above characterized range (>300 g/L resin)",
        eff="Reduced HCP removal capacity (viral clearance robust)", ctl="Load by mass balance; design space 50-300 g/L; viral clearance validated to 300 g/L"),
    ("aex", "flow"): dict(cqas=["lrv_xmulv", "lrv_mvm"], fm="Operating flow rate above 450 cm/hr",
        eff="Reduced residence time / viral clearance margin", ctl="Flow controlled by skid; viral clearance validated <= 450 cm/hr"),
    ("virus_filtration", "filtration_volume"): dict(cqas=["lrv_mvm"], fm="Volumetric load above 105 L/m2",
        eff="Reduced MVM log-reduction (parvovirus breakthrough risk)", ctl="Volumetric load limited to <= 105 L/m2 incl. chase; filter integrity test"),
    ("virus_filtration", "pressure"): dict(cqas=["lrv_mvm", "lrv_xmulv"], fm="Pressure outside filter operating limits",
        eff="Potential impact on virus removal at high load", ctl="Pressure within manufacturer limits; monitored; pre/post integrity test"),
    ("harvest", "centrifuge_g"): dict(cqas=["performance"], fm="Centrifugation rcf outside 6000-12000 g",
        eff="Affects clarification/yield (no CQA impact)", ctl="Centrifuge speed controlled; turbidity IPC", kpp=True),
    ("harvest", "depth_filter_load"): dict(cqas=["performance"], fm="Depth filter load outside 80-160 L/m2",
        eff="Affects filter capacity/turbidity (no CQA impact)", ctl="Sizing by area; turbidity IPC", kpp=True),
    ("harvest", "turbidity"): dict(cqas=["performance"], fm="Post-clarification turbidity high",
        eff="Feed quality to Protein A (no CQA impact)", ctl="Turbidity IPC; Protein A robust to feedstock", gpp=True),
    ("ufdf", "diavolumes"): dict(cqas=["performance"], fm="Diavolumes < 7",
        eff="Incomplete buffer exchange (formulation; not part of DS characterization)", ctl="Diavolume count controlled; buffer exchange IPC", kpp=True),
    ("ufdf", "tmp"): dict(cqas=["performance"], fm="TMP outside 10-15 psi",
        eff="Affects flux/processing (formulation)", ctl="TMP controlled by skid", kpp=True),
    ("ufdf", "final_conc"): dict(cqas=["performance"], fm="Final concentration outside 65-85 g/L",
        eff="DS concentration spec (formulation)", ctl="Concentration by UV/mass; release test", kpp=True),
}


def severity(cqas):
    return max(CQA_SEVERITY.get(c, 4) for c in cqas)


def detection_res(cqas):
    return max(CQA_DET_RES.get(c, 4) for c in cqas)


def build_rows(cfg):
    rows = []
    for key in cfg.train_order:
        uo = cfg.unit_op(key)
        for p in uo.parameters:
            c = CONTENT.get((key, p.key))
            if not c:
                c = dict(cqas=["performance"], fm="Parameter outside characterized range",
                         eff="No significant product-quality impact demonstrated", ctl="Standard equipment control", gpp=True)
            cqas = c["cqas"]
            S = severity(cqas)
            quality = cqas != ["performance"]
            # initial (pre-characterization) occurrence/detection: uncertain effect, limited IPC
            O_init = 7 if quality else 4
            D_init = 10 if ("lrv_xmulv" in cqas or "lrv_mvm" in cqas) else (8 if quality else 6)
            # residual (post-characterization) occurrence/detection: proven ranges + control strategy
            O_res = 4 if c.get("cpp") else (2 if quality else 3)
            D_res = detection_res(cqas)
            rpn_i, rpn_r = S * O_init * D_init, S * O_res * D_res
            classification = p.classification
            # CPP status is set by the PRE-mitigation (initial) risk of impacting a CQA:
            # a parameter that impacts a CQA with Severity >= 8 or initial RPN > 72 is a CPP.
            # The control strategy then determines CPP vs well-controlled CPP (WC-CPP).
            cpp_sev, cpp_rpn = cfg.risk["thresholds"]["cpp_severity"], cfg.risk["thresholds"]["cpp_rpn"]
            is_cpp = quality and ((S >= cpp_sev) or (rpn_i > cpp_rpn))
            # Residual risk band follows the objective residual RPN (same thresholds as the
            # RPN colouring). A high-severity attribute retains an elevated RPN even when well
            # controlled — which is precisely why it is designated (WC-)CPP and enhanced-controlled.
            residual = "High" if rpn_r > cpp_rpn else ("Medium" if rpn_r > cfg.risk["thresholds"]["medium_rpn"] else "Low")
            rows.append({
                "Step": uo.step, "Unit Operation": uo.name, "Process Parameter": p.name,
                "Set-point": p.setpoint, "NOR": f"{p.nor[0]}-{p.nor[1]}", "PAR (characterized)": f"{p.par[0]}-{p.par[1]}",
                "Potential Failure Mode": c["fm"], "Potential Effect (CQA / performance)": c["eff"],
                "CQA(s) impacted": ", ".join("process performance" if x == "performance" else x for x in cqas),
                "S": S, "O (init)": O_init, "D (init)": D_init, "RPN (init)": rpn_i,
                "O (res)": O_res, "D (res)": D_res, "RPN (residual)": rpn_r,
                "CPP rule (S>=8 or RPN>72)": "CPP" if is_cpp else "not CPP",
                "Designation": classification, "Control Strategy": c["ctl"], "Residual Risk": residual,
            })
    return pd.DataFrame(rows)


# --- workbook styling helpers -------------------------------------------------
def style_header(ws, row, ncol):
    for j in range(1, ncol + 1):
        cell = ws.cell(row=row, column=j)
        cell.fill = HFILL; cell.font = HFONT; cell.alignment = CTR; cell.border = BORDER


def rpn_fill(v):
    if v > 72:
        return PatternFill("solid", fgColor=RED)
    if v > 48:
        return PatternFill("solid", fgColor=AMBER)
    return PatternFill("solid", fgColor=GREEN)


def risk_fill(v):
    return {"High": PatternFill("solid", fgColor=RED), "Medium": PatternFill("solid", fgColor=AMBER),
            "Low": PatternFill("solid", fgColor=GREEN)}.get(v, PatternFill("solid", fgColor=WHITE))


def main():
    cfg = load_config()
    df = build_rows(cfg)
    wb = Workbook()

    # ---------- Cover ----------
    ws = wb.active; ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "A-Mab Drug Substance"; ws["B2"].font = Font(size=20, bold=True, color=NAVY)
    ws["B3"] = "Post-Characterization Process Risk Assessment (FMEA)"; ws["B3"].font = Font(size=14, bold=True, color=BLUE)
    meta = [
        ("Product", f"{cfg.meta['product']} — {cfg.meta['modality']}"),
        ("Document type", "Post-Process-Characterization Quality Risk Assessment (Risk Assessment #3/#4)"),
        ("Methodology", cfg.risk["method"]),
        ("CPP decision rule", cfg.risk["cpp_rule"]),
        ("Residual-risk note", "The residual-RPN band (>72 High, 48–72 Medium, ≤48 Low) is objective and reflects post-characterization Occurrence/Detection. A high-severity CQA (e.g. viral safety, glycosylation) retains an elevated RPN even when well controlled — which is precisely why it is designated (WC-)CPP and subject to enhanced control. Acceptability is governed by the control strategy, not the RPN band alone."),
        ("Scope", "Drug-substance train Steps 3-10 (bioreactor through UF/DF); process parameters only"),
        ("Basis", "A-Mab Case Study v2.1 (CMC Biotech Working Group, 2009); ICH Q8/Q9/Q11; FDA 2011"),
        ("Model seed", str(cfg.seed)),
        ("Parameters assessed", str(len(df))),
        ("CPPs / WC-CPPs", str(int(df["Designation"].isin(["CPP", "WC-CPP"]).sum()))),
    ]
    r = 5
    for k, v in meta:
        ws.cell(r, 2, k).font = Font(bold=True, color=NAVY)
        ws.cell(r, 3, v).alignment = WRAP
        r += 1
    ws.cell(r + 1, 2, "Sheets: CQA Criticality · Process FMEA (Post-PC) · Scoring Scales · Summary").font = Font(italic=True, color="808080")
    ws.column_dimensions["B"].width = 22; ws.column_dimensions["C"].width = 95

    # ---------- CQA Criticality ----------
    ws = wb.create_sheet("CQA Criticality")
    cqa_cols = ["CQA", "Category", "Acceptance", "Clinical range", "Criticality", "Tool#1 Impact",
                "Tool#1 Uncertainty", "Tool#1 Score", "Tool#2 Severity", "Set primarily by"]
    ws.append(["A-Mab Critical Quality Attributes — criticality (Tool #1: Impact x Uncertainty; H/VH = Critical)"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cqa_cols))
    ws.cell(1, 1).font = Font(bold=True, size=12, color=NAVY)
    ws.append(cqa_cols); style_header(ws, 2, len(cqa_cols))
    for c in cfg.cqas:
        ws.append([c["name"], c["category"], f"{c['acceptance'][0]}-{c['acceptance'][1]} {c['unit']}",
                   f"{c.get('clinical',[None,None])[0]}-{c.get('clinical',[None,None])[1]}", c["criticality_level"],
                   c.get("tool1_impact"), c.get("tool1_uncertainty"), c.get("tool1_score"),
                   c.get("tool2_severity"), c.get("set_by")])
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=len(cqa_cols)):
        for cell in row:
            cell.border = BORDER; cell.alignment = WRAP
        lvl = row[4].value
        if lvl in ("VH", "H", "M-H"):
            row[4].fill = PatternFill("solid", fgColor=RED if lvl == "VH" else AMBER)
    widths = [30, 20, 20, 14, 10, 12, 14, 11, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    # ---------- Process FMEA ----------
    ws = wb.create_sheet("Process FMEA (Post-PC)")
    cols = list(df.columns)
    ws.append([f"A-Mab Post-Characterization Process FMEA — {cfg.risk['method']}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.cell(1, 1).font = Font(bold=True, size=12, color=NAVY)
    ws.append(cols); style_header(ws, 2, len(cols))
    for _, rrow in df.iterrows():
        ws.append([rrow[c] for c in cols])
    idx = {c: i + 1 for i, c in enumerate(cols)}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=len(cols)):
        for cell in row:
            cell.border = BORDER; cell.alignment = WRAP
        row[idx["RPN (init)"] - 1].fill = rpn_fill(row[idx["RPN (init)"] - 1].value)
        row[idx["RPN (residual)"] - 1].fill = rpn_fill(row[idx["RPN (residual)"] - 1].value)
        row[idx["Residual Risk"] - 1].fill = risk_fill(row[idx["Residual Risk"] - 1].value)
        des = row[idx["Designation"] - 1]
        des.fill = {"CPP": PatternFill("solid", fgColor=RED), "WC-CPP": PatternFill("solid", fgColor=AMBER),
                    "KPP": PatternFill("solid", fgColor=LIGHT), "GPP": PatternFill("solid", fgColor=WHITE)}.get(des.value, PatternFill("solid", fgColor=WHITE))
        des.font = Font(bold=True)
    # column widths
    wmap = {"Unit Operation": 20, "Process Parameter": 20, "Potential Failure Mode": 34,
            "Potential Effect (CQA / performance)": 34, "CQA(s) impacted": 20, "Control Strategy": 40,
            "NOR": 12, "PAR (characterized)": 16, "Designation": 12, "Residual Risk": 11,
            "CPP rule (S>=8 or RPN>72)": 14}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = wmap.get(c, 8 if len(c) <= 3 else 12)
    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{ws.max_row}"
    dv = DataValidation(type="list", formula1='"CPP,WC-CPP,KPP,GPP,non-CPP"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(idx['Designation'])}3:{get_column_letter(idx['Designation'])}{ws.max_row}")

    # ---------- Scoring Scales ----------
    ws = wb.create_sheet("Scoring Scales")
    r = 1
    for title, scale in [("Severity (S) — Table 5.22", cfg.risk["severity_scale"]),
                         ("Occurrence (O) — Table 5.23", cfg.risk["occurrence_scale"]),
                         ("Detection (D) — Table 5.24", cfg.risk["detection_scale"])]:
        ws.cell(r, 1, title).font = Font(bold=True, size=12, color=NAVY); r += 1
        ws.cell(r, 1, "Score"); ws.cell(r, 2, "Label"); ws.cell(r, 3, "Definition")
        style_header(ws, r, 3); r += 1
        for item in scale:
            ws.cell(r, 1, item["score"]); ws.cell(r, 2, item["label"]); ws.cell(r, 3, item["meaning"])
            for j in (1, 2, 3):
                ws.cell(r, j).border = BORDER; ws.cell(r, j).alignment = WRAP
            r += 1
        r += 1
    ws.cell(r, 1, cfg.risk["cpp_rule"]).font = Font(italic=True, bold=True, color=BLUE)
    ws.column_dimensions["A"].width = 8; ws.column_dimensions["B"].width = 16; ws.column_dimensions["C"].width = 90

    # ---------- Summary ----------
    ws = wb.create_sheet("Summary")
    ws.cell(1, 1, "Risk-assessment summary").font = Font(bold=True, size=13, color=NAVY)
    counts = df["Designation"].value_counts()
    ws.cell(3, 1, "Parameter classification").font = Font(bold=True)
    ws.cell(3, 2, "Count").font = Font(bold=True)
    r = 4
    for cls in ["CPP", "WC-CPP", "KPP", "GPP", "non-CPP"]:
        ws.cell(r, 1, cls); ws.cell(r, 2, int(counts.get(cls, 0))); r += 1
    ws.cell(r, 1, "Total"); ws.cell(r, 2, len(df)); ws.cell(r, 1).font = Font(bold=True); r += 2
    med_i = df["RPN (init)"].median(); med_r = df["RPN (residual)"].median()
    for lab, val in [("Median RPN before characterization", f"{med_i:.0f}"),
                     ("Median RPN after characterization + controls", f"{med_r:.0f}"),
                     ("Median RPN reduction", f"{(1-med_r/med_i)*100:.0f}%"),
                     ("Parameters that are CPP or WC-CPP", int(df["Designation"].isin(["CPP", "WC-CPP"]).sum())),
                     ("High residual-risk parameters (RPN_res > 72)", int((df["RPN (residual)"] > 72).sum()))]:
        ws.cell(r, 1, lab).font = Font(bold=True); ws.cell(r, 2, val); r += 1
    ws.column_dimensions["A"].width = 46; ws.column_dimensions["B"].width = 16

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"wrote {os.path.relpath(OUT, ROOT)}")
    print(f"  {len(df)} parameters | CPP/WC-CPP={int(df['Designation'].isin(['CPP','WC-CPP']).sum())} | "
          f"median RPN {med_i:.0f} -> {med_r:.0f}")


if __name__ == "__main__":
    main()
